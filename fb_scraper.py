import os
import time
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import Workbook
from openpyxl.styles import Font

#CHANGE MO NG EMAIL AND PASSWORD
EMAIL = '####'  
PASSWORD = '####'        

def create_browser():
    options = Options()
    options.add_argument("--disable-notifications")
    options.add_argument("--disable-infobars")
    options.add_argument("--mute-audio")
    browser = webdriver.Chrome(options=options)
    return browser

def fb_login(browser):
    browser.get("https://facebook.com/")
    print("[*] Logging into Facebook...")

    browser.find_element(By.NAME, 'email').send_keys(EMAIL)
    browser.find_element(By.NAME, 'pass').send_keys(PASSWORD)
    browser.find_element(By.NAME, 'login').click()
    time.sleep(3)  

def scroll_to_bottom(browser):
    print("[*] Scrolling through the friends list...")
    try:
        previous_count = 0
        timeout = 300  
        start_time = time.time()

        while True:
            browser.execute_script("window.scrollBy(0, 1000);")
            time.sleep(2)
            current_count = len(browser.find_elements(By.CSS_SELECTOR, 
                "span.x193iq5w.xeuugli.x13faqbe.x1vvkbs.x1lliihq.x1s928wv.xhkezso.x1gmr53x.x1cpjm7i.x1fgarty.x1943h6x.xudqn12.x676frb.x1lkfr7t.x1lbecb7.x1s688f.xzsf02u"))
            print(f"[DEBUG] Loaded friends count: {current_count}")

            if current_count == previous_count or time.time() - start_time > timeout:
                print("[*] Stopped scrolling.")
                break

            previous_count = current_count

    except Exception as e:
        print(f"[!] Error during scrolling: {e}")


def scan_friends(browser):
    print('[*] Scanning the friends list...')
    friends = []

    friend_elements = browser.find_elements(By.CSS_SELECTOR, 
        "span.x193iq5w.xeuugli.x13faqbe.x1vvkbs.x1lliihq.x1s928wv.xhkezso.x1gmr53x.x1cpjm7i.x1fgarty.x1943h6x.xudqn12.x676frb.x1lkfr7t.x1lbecb7.x1s688f.xzsf02u")
    print(f"[DEBUG] Found {len(friend_elements)} elements.")

    for friend in friend_elements:
        friend_name = friend.text
        try:
            friend_url = friend.find_element(By.XPATH, "./ancestor::a").get_attribute("href")
        except:
            friend_url = None

        if friend_name:
            friends.append({
                'name': friend_name,
                'url': friend_url if friend_url else "N/A"
            })
    print(f"[+] Found {len(friends)} friends!")
    return friends


def save_to_excel(friends, filename="friends_list.xlsx"):
    wb = Workbook()
    ws = wb.active
    ws.title = "Friends List"


    bold_font = Font(bold=True)
    ws["A1"] = "Name"
    ws["B1"] = "URL"
    ws["A1"].font = bold_font
    ws["B1"].font = bold_font

    
    for index, friend in enumerate(friends, start=2):
        ws[f"A{index}"] = friend['name']
        ws[f"B{index}"] = friend['url']

    
    ws.column_dimensions["A"].width = 50  
    ws.column_dimensions["B"].width = 80  

    wb.save(filename)
    print(f"[+] Friends data saved to {filename}")


def main():
    browser = create_browser()
    
    try:
        fb_login(browser)
        time.sleep(3)  
        
        browser.get("####") #CHANGE MO NG PROFILE AND FRIENDS LINK (https://www.facebook.com/profile.php?id=1000*******1337&sk=friends) 
        time.sleep(2) 
        
        scroll_to_bottom(browser)  
        
        friends = scan_friends(browser) 
        
        save_to_excel(friends, "friends_list.xlsx") 
        
    except Exception as e:
        print(f"[!] Error: {e}")
    
    finally:
        browser.quit()

if __name__ == "__main__":
    main()
